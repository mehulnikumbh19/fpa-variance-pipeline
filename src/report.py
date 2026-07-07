"""Excel report writer: formatted variance workbook via openpyxl.

Four tabs: Executive Summary (findings first), Variance Detail (full table
with conditional formatting), Trend View (cost center x month heat grid),
and Data Quality (the validation evidence — the report shows its own
controls so a reviewer can trust the numbers before reading them).
Formatting is deliberately restrained: banker-grey headers, standard
red/green/amber flag colors, nothing else.
"""

import calendar
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FMT_USD = "$#,##0"
FMT_PCT = "0.0%"
HEADER_FILL = PatternFill("solid", start_color="1F3864")  # navy
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, color="808080")
# Classic Excel conditional-format palette — familiar to any spreadsheet
# reviewer: red = unfavorable flag, green = favorable flag, amber = Watch.
FILL_RED = PatternFill("solid", start_color="FFC7CE")
FILL_GREEN = PatternFill("solid", start_color="C6EFCE")
FILL_AMBER = PatternFill("solid", start_color="FFEB9C")

DETAIL_COLS = [("month", "Month", None), ("cost_center", "Cost Center", None),
               ("account_group", "Account Group", None),
               ("budget", "Budget", FMT_USD), ("actual", "Actual", FMT_USD),
               ("variance_usd", "Variance ($)", FMT_USD),
               ("variance_pct", "Variance (%)", FMT_PCT),
               ("direction", "Direction", None),
               ("severity", "Severity", None)]


def _severity_fill(direction: str, severity: str) -> PatternFill | None:
    if not severity:
        return None
    if severity == "Watch":
        return FILL_AMBER
    return FILL_RED if direction == "Unfavorable" else FILL_GREEN


def _write_table(ws, df: pd.DataFrame, cols, start_row: int) -> int:
    """Write a header + data block; returns the last row used."""
    for j, (_, header, _) in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=j, value=header)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        vals = dict(zip(df.columns, row))
        for j, (key, _, fmt) in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=j, value=vals[key])
            if fmt:
                cell.number_format = fmt
            if key in ("variance_usd", "variance_pct", "severity"):
                fill = _severity_fill(vals.get("direction", ""),
                                      vals.get("severity", ""))
                if fill:
                    cell.fill = fill
    return start_row + len(df)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _narratives(ytd_cc: pd.DataFrame, ytd_ag: pd.DataFrame,
                detail: pd.DataFrame) -> list[str]:
    """One line per escalated cost center, naming the primary driver.

    'Escalated' = the cost center carries an Escalate-tier flag at any
    grain (monthly or YTD) — a monthly blowout deserves an executive
    sentence even when twelve months of denominator dilute the YTD %.
    """
    esc = set(detail.loc[detail["severity"] == "Escalate", "cost_center"]) \
        | set(ytd_ag.loc[ytd_ag["severity"] == "Escalate", "cost_center"])
    lines = []
    ranked = ytd_cc[ytd_cc["cost_center"].isin(esc)] \
        .sort_values("variance_usd", key=abs, ascending=False)
    for row in ranked.itertuples(index=False):
        drivers = ytd_ag[(ytd_ag["cost_center"] == row.cost_center)
                         & (ytd_ag["variance_usd"] * row.variance_usd > 0)]
        driver = (drivers.iloc[0]["account_group"]
                  if len(drivers) else "multiple account groups")
        lines.append(
            f"{row.cost_center} is ${abs(row.variance_usd):,.0f} "
            f"({abs(row.variance_pct):.1%}) {row.direction.lower()} YTD, "
            f"driven primarily by {driver}.")
    return lines


def _exec_summary(ws, detail, ytd_cc, ytd_ag, trends):
    ws["A1"] = "FY2026 Budget vs. Actual Variance Report — Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Fictional community bank; all data is synthetic. Generated "
                + datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws["A2"].font = NOTE_FONT

    row = 4
    ws.cell(row=row, column=1, value="Monthly flags by severity").font = \
        Font(bold=True)
    counts = detail.loc[detail["severity"] != "", "severity"].value_counts()
    for tier in ("Escalate", "Review", "Watch"):
        row += 1
        ws.cell(row=row, column=1, value=tier)
        ws.cell(row=row, column=2, value=int(counts.get(tier, 0)))

    row += 2
    ws.cell(row=row, column=1,
            value="Top 10 YTD variances by absolute dollar impact").font = \
        Font(bold=True)
    top10_cols = [("cost_center", "Cost Center", None),
                  ("account_group", "Account Group", None),
                  ("budget", "Budget YTD", FMT_USD),
                  ("actual", "Actual YTD", FMT_USD),
                  ("variance_usd", "Variance ($)", FMT_USD),
                  ("variance_pct", "Variance (%)", FMT_PCT),
                  ("direction", "Direction", None),
                  ("severity", "Severity", None)]
    row = _write_table(ws, ytd_ag.head(10), top10_cols, row + 1)

    row += 2
    ws.cell(row=row, column=1, value="Escalated items — narrative").font = \
        Font(bold=True)
    for line in _narratives(ytd_cc, ytd_ag, detail):
        row += 1
        ws.cell(row=row, column=1, value="• " + line)

    row += 2
    ws.cell(row=row, column=1,
            value="Trend watch — 3+ consecutive unfavorable months").font = \
        Font(bold=True)
    if trends.empty:
        row += 1
        ws.cell(row=row, column=1, value="None identified.")
    for t in trends.itertuples(index=False):
        row += 1
        start = calendar.month_abbr[t.run_start_month]
        end = calendar.month_abbr[t.run_start_month + t.run_length - 1]
        ws.cell(row=row, column=1,
                value=f"• {t.cost_center}: {t.run_length} consecutive "
                      f"unfavorable months ({start}–{end})")
    _autosize(ws, [34, 18, 14, 14, 14, 13, 13, 11])


def _detail_tab(ws, detail):
    end = _write_table(ws, detail[[c for c, _, _ in DETAIL_COLS]],
                       DETAIL_COLS, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLS))}{end}"
    _autosize(ws, [8, 28, 24, 12, 12, 13, 13, 13, 10])


def _trend_tab(ws, detail):
    # Cost center x month grid of variance % at cost-center total grain.
    totals = detail.groupby(["cost_center", "month"]).sum(numeric_only=True)
    totals["variance_pct"] = ((totals["actual"] - totals["budget"])
                              / totals["budget"])
    grid = totals["variance_pct"].unstack("month")

    ws.cell(row=1, column=1, value="Cost Center")
    for m in grid.columns:
        ws.cell(row=1, column=m + 1, value=calendar.month_abbr[m])
    for cell in ws[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for i, (cc, vals) in enumerate(grid.iterrows(), start=2):
        ws.cell(row=i, column=1, value=cc)
        for m, v in vals.items():
            c = ws.cell(row=i, column=m + 1, value=round(float(v), 4))
            c.number_format = FMT_PCT
    # Diverging scale: green = under budget (favorable), red = over.
    rng = f"B2:M{len(grid) + 1}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=-0.15, start_color="63BE7B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=0.15, end_color="F8696B"))
    ws.freeze_panes = "B2"
    _autosize(ws, [30] + [9] * 12)


def _quality_tab(ws, validation_results):
    ws["A1"] = "Input validation results (see output/validation_log.txt)"
    ws["A1"].font = TITLE_FONT
    headers = ["Check", "Control Objective", "Result", "Detail"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=j, value=h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for i, r in enumerate(validation_results, start=4):
        ws.cell(row=i, column=1, value=r.name)
        ws.cell(row=i, column=2, value=r.objective)
        res = ws.cell(row=i, column=3, value="PASS" if r.passed else "FAIL")
        res.fill = FILL_GREEN if r.passed else FILL_RED
        detail_cell = ws.cell(row=i, column=4, value=r.detail)
        detail_cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A4"
    _autosize(ws, [24, 20, 10, 70])


def write_report(path, detail, ytd_cc, ytd_ag, trends, validation_results):
    """Assemble and save the four-tab workbook."""
    wb = Workbook()
    _exec_summary(wb.active, detail, ytd_cc, ytd_ag, trends)
    wb.active.title = "Executive Summary"
    _detail_tab(wb.create_sheet("Variance Detail"), detail)
    _trend_tab(wb.create_sheet("Trend View"), detail)
    _quality_tab(wb.create_sheet("Data Quality"), validation_results)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
